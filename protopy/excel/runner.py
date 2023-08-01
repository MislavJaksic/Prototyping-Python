import openpyxl

def unmerge_and_fill_cells(input_file, output_file):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    # Select the first sheet (you can modify this if needed)
    sheet = wb.active

    # Make a copy of the merged cell ranges to avoid the 'set changed size during iteration' error
    merged_ranges_copy = sheet.merged_cells.ranges.copy()

    # Unmerge all cells in the copied ranges and fill with data from the top-left cell
    for merged_range in merged_ranges_copy:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        top_left_cell = sheet.cell(min_row, min_col)
        value_to_fill = top_left_cell.value
        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row, col)
                cell.value = value_to_fill

    # Save the unmerged data to a new file
    wb.save(output_file)

if __name__ == "__main__":
    # Replace 'input_file.xlsx' and 'output_file.xlsx' with your desired file names.
    input_file = "input_file.xlsx"
    output_file = "output_file.xlsx"
    unmerge_and_fill_cells(input_file, output_file)
    print("Unmerged cells and filled with data. Saved to", output_file)
